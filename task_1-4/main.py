import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt

wb = load_workbook(r'c:\Users\Антон\Downloads\students.xlsx')
grades = wb["Оценки"] #лист с оценками
math = []
fizika = []
info = []
for row in grades.iter_rows(min_row=2, values_only=True):
    math.append(row[1])
    fizika.append(row[2])
    info.append(row[3])
def calc(x):
    return (sum(x)/len(x))
avg_m = calc(math)
avg_f = calc(fizika)
avg_i = calc(info)
print(f"Cредний балл по математике = {avg_m:.2f}")
print(f"Cредний балл по физике = {avg_f:.2f}")
print(f"Cредний балл по информатике = {avg_i:.2f}")

# Текст таблиц и используемые перменные
предметы = ['Математика', 'Физика', 'Информатика']
средние = [avg_m, avg_f, avg_i]


# Столбчатая диаграмма (самый простой график)
plt.figure(figsize=(8, 5))  # размер графика
plt.bar(предметы, средние, color=['blue', 'green', 'red'])

# Настройки графика
plt.title('Средние оценки', fontsize=14)
plt.ylabel('Средний балл', fontsize=12)
plt.ylim(0, 10)  # от 0 до 10 (для 10-балльной системы)

# Добавляем значения на столбцы
for i, v in enumerate(средние):
    plt.text(i, v + 0.1, f'{v:.2f}', ha='center', fontsize=11)

plt.grid(axis='y', alpha=0.3, linestyle='--')
plt.show()